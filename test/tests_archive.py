import openpyxl
from pypdf import PdfReader
import csv
from zipfile import ZipFile
from module.path import *


def test_read_pdf_file():
    with ZipFile(ARCHIVE, 'r') as archive:
        with archive.open('pdf.pdf') as file_in_zip:
            pdf_reader = PdfReader(file_in_zip)
            assert len(pdf_reader.pages) > 0


def test_read_csv_file():
    with ZipFile(ARCHIVE, 'r') as archive:
        with archive.open('csv.csv') as file_in_zip:
            decoded_content = file_in_zip.read().decode('utf-8')
            csv_reader = csv.reader(decoded_content.splitlines())
            rows = list(csv_reader)
            assert len(rows) > 0


def test_csv_content():
    expected_first_word = 'OrderID*'
    with ZipFile(ARCHIVE, "r") as extractor:
        extractor.extract("csv.csv", RESOURCES)
    with open(EXTRACTED_CSV, "r", encoding="UTF-8") as file:
        csv_reader = csv.reader(file)
        header_row = next(csv_reader, None)
        if header_row:
            first_word = header_row[0]
            assert first_word == expected_first_word


def test_read_xlsx_content():
    with ZipFile(ARCHIVE, 'r') as archive:
        with archive.open('xlsx.xlsx') as file_in_zip:
            workbook = openpyxl.load_workbook(file_in_zip)
            sheet = workbook.active
            assert sheet.max_row > 0 and sheet.max_column > 0
